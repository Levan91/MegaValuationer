import os
import re
import json
import sys
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import openai
import time  # Add at the top if not present

# Setup OpenAI API key and GPT enablement
try:
    with open("/Users/levan/Documents/LIVELISTINGS/openai_key.txt", "r") as f:
        openai.api_key = f.read().strip()
    gpt_enabled = True
    print("✅ GPT enabled using API key from file.")
except Exception as e:
    gpt_enabled = False
    print(f"⚠️ GPT is disabled. Could not read API key: {e}")

def ai_confirm_plot_size(candidates, title, description):
    # ... existing code ...
    return ""

# Determine project from command-line argument (default to 'Maple')
project = sys.argv[1] if len(sys.argv) > 1 else "Maple"
HTML_FOLDER = os.path.join("/Users/levan/Documents/LIVELISTINGS", project)
BASE_NAME = project
LAYOUTS_FILE = os.path.join("/Users/levan/Documents/MegaValuationer/Data/Layout Types", f"{BASE_NAME}_layouts.xlsx")
OUTPUT_FILE = os.path.join("/Users/levan/Documents/MegaValuationer/Data/Listings", f"{BASE_NAME}_listings.xlsx")
print(f"[DEBUG] Listings will be saved to: {OUTPUT_FILE}")
DATE_FORMAT = "%Y-%m-%dT%H:%M:%SZ"

for old_ref in sorted(old_refs):
    old_url = old[old_ref].get("URL")
    # Exact match by reference
    if old_ref in new:
        rec = new[old_ref].copy()
        # --- Preserve manual edits for key fields ---
        fields_to_preserve = ["Unit Type", "Unit Number", "Type"]
        for field in fields_to_preserve:
            if old[old_ref].get(field):
                rec[field] = old[old_ref][field]
        rec["Availability"] = "Available"
        rec["_color"] = None
        merged.append((old_ref, rec))
        processed_new_refs.add(old_ref)
    # Fallback: URL match with a different new reference
    elif old_url in url_to_ref:
        new_ref = url_to_ref[old_url]
        rec = new[new_ref].copy()
        # --- Preserve manual edits for key fields ---
        fields_to_preserve = ["Unit Type", "Unit Number", "Type"]
        for field in fields_to_preserve:
            if old[old_ref].get(field):
                rec[field] = old[old_ref][field]
        rec["Availability"] = "Available"
        rec["_color"] = None
        # Preserve the original reference code
        rec["Reference Number"] = old_ref
        merged.append((old_ref, rec))
        processed_new_refs.add(new_ref)
    else:
        rec = old[old_ref].copy()
        rec["Availability"] = "Not available"
        rec["_color"] = "removed"
        merged.append((old_ref, rec)) 